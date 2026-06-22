"""Create a visual preview of the generated Excel workbook."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


REPORTS = Path("reports")
WORKBOOK = REPORTS / "final_report.xlsx"
PREVIEW = REPORTS / "excel_workbook_preview.png"
ASSET_COPY = Path("assets/ch10/web/excel_workbook_preview.png")


def font(size: int, bold: bool = False):
    candidates = [
        Path("C:/Windows/Fonts/msyhbd.ttc") if bold else Path("C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/simhei.ttf"),
        Path("C:/Windows/Fonts/arial.ttf"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size=size)
    return ImageFont.load_default()


def read_cells() -> list[list[str]]:
    if not WORKBOOK.exists():
        raise FileNotFoundError("Run 04_generate_office_pack.py first.")
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(min_row=1, max_row=6, max_col=4, values_only=True):
        rows.append(["" if value is None else str(value) for value in row])
    return rows


def write_preview(rows: list[list[str]]) -> None:
    image = Image.new("RGB", (1500, 980), "#F7F8FB")
    draw = ImageDraw.Draw(image)
    draw.rounded_rectangle((90, 70, 1410, 900), radius=26, fill="#FFFFFF", outline="#D8E0EC", width=3)
    draw.text((150, 125), "Excel Workbook Preview", fill="#162033", font=font(50, True))
    draw.text((150, 198), "Python 读取 final_report.xlsx，生成可检查的表格预览。", fill="#5F6673", font=font(27))

    x0, y0 = 150, 300
    widths = [210, 210, 470, 170]
    row_h = 70
    for r, row in enumerate(rows):
        x = x0
        for c, value in enumerate(row):
            fill = "#2F6BFF" if r == 0 else "#F1F5F9"
            text_fill = "#FFFFFF" if r == 0 else "#162033"
            draw.rounded_rectangle((x, y0 + r * row_h, x + widths[c], y0 + r * row_h + 56), radius=12, fill=fill, outline="#E2E8F0", width=2)
            draw.text((x + 18, y0 + r * row_h + 14), value, fill=text_fill, font=font(23, r == 0))
            x += widths[c] + 12

    draw.rounded_rectangle((250, 840, 1250, 895), radius=20, fill="#EEF6FF", outline="#9CC8FF", width=2)
    draw.text((320, 854), "Preview source: reports/final_report.xlsx", fill="#28517A", font=font(22))

    PREVIEW.parent.mkdir(exist_ok=True)
    image.save(PREVIEW, optimize=True, quality=95)
    ASSET_COPY.parent.mkdir(parents=True, exist_ok=True)
    image.save(ASSET_COPY, optimize=True, quality=95)


def main() -> None:
    rows = read_cells()
    write_preview(rows)
    print(f"已生成 {PREVIEW}")
    print(f"已复制到 {ASSET_COPY}")


if __name__ == "__main__":
    main()
